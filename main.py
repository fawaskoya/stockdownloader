from flask import Flask, render_template, request, send_file, session
import yfinance as yf
import pandas as pd
import openpyxl
import io
import os
from datetime import datetime
import json

app = Flask(__name__)
app.secret_key = os.urandom(24)  # Required for session

def get_ticker_suggestions(query):
    url = f"https://query1.finance.yahoo.com/v1/finance/search?q={query}"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        if 'quotes' in data:
            return [{'symbol': quote['symbol'], 'name': quote['longname']} 
                   for quote in data['quotes'] if 'symbol' in quote and 'longname' in quote]
    return []

def get_financial_ratios(ticker):
    stock = yf.Ticker(ticker)
    info = stock.info
    
    ratios = {
        'Valuation': {
            'P/E (TTM)': info.get('trailingPE', 'N/A'),
            'Forward P/E': info.get('forwardPE', 'N/A'),
            'PEG Ratio': info.get('pegRatio', 'N/A'),
            'Price/Book': info.get('priceToBook', 'N/A')
        },
        'Profitability': {
            'ROE': info.get('returnOnEquity', 'N/A'),
            'ROA': info.get('returnOnAssets', 'N/A'),
            'Operating Margin': info.get('operatingMargins', 'N/A'),
            'Profit Margin': info.get('profitMargins', 'N/A')
        },
        'Solvency': {
            'Debt/Equity': info.get('debtToEquity', 'N/A'),
            'Book Value': info.get('bookValue', 'N/A')
        },
        'Liquidity': {
            'Current Ratio': info.get('currentRatio', 'N/A')
        },
        'Dividend': {
            'Dividend Yield': info.get('dividendYield', 'N/A'),
            'Payout Ratio': info.get('payoutRatio', 'N/A')
        }
    }
    return ratios

def calculate_valuation_models(ticker, required_return=0.1, growth_rate=0.05):
    stock = yf.Ticker(ticker)
    info = stock.info
    
    # Dividend Discount Model
    current_dividend = info.get('dividendRate', 0)
    if current_dividend and required_return > growth_rate:
        ddm_value = current_dividend / (required_return - growth_rate)
    else:
        ddm_value = 'N/A'
    
    # Earnings-Based Valuation
    eps = info.get('trailingEps', 0)
    pe_ratio = info.get('trailingPE', 0)
    if eps and pe_ratio:
        earnings_value = eps * pe_ratio
    else:
        earnings_value = 'N/A'
    
    # Book Value-Based Valuation
    book_value = info.get('bookValue', 0)
    pb_ratio = info.get('priceToBook', 0)
    if book_value and pb_ratio:
        book_value_valuation = book_value * pb_ratio
    else:
        book_value_valuation = 'N/A'
    
    return {
        'DDM': ddm_value,
        'Earnings-Based': earnings_value,
        'Book Value-Based': book_value_valuation
    }

def format_excel_worksheet(worksheet, title):
    # Set column widths
    for col in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 20

    # Add header
    header_fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Add title
    worksheet.insert_rows(1)
    worksheet.merge_cells(f'A1:{get_column_letter(worksheet.max_column)}1')
    title_cell = worksheet['A1']
    title_cell.value = title
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

def calculate_technical_indicators(data):
    """Calculate technical indicators for the stock data using the 'ta' library"""
    indicators = {}
    # Ensure these are Series, not DataFrames
    close = data['Close'] if isinstance(data['Close'], pd.Series) else data['Close'].squeeze()
    volume = data['Volume'] if isinstance(data['Volume'], pd.Series) else data['Volume'].squeeze()

    # Moving Averages
    indicators['SMA_20'] = ta.trend.sma_indicator(close, window=20)
    indicators['SMA_50'] = ta.trend.sma_indicator(close, window=50)
    indicators['SMA_200'] = ta.trend.sma_indicator(close, window=200)

    # RSI
    indicators['RSI'] = ta.momentum.rsi(close, window=14)

    # MACD
    macd = ta.trend.macd(close)
    macd_signal = ta.trend.macd_signal(close)
    macd_diff = ta.trend.macd_diff(close)
    indicators['MACD'] = macd
    indicators['MACD_Signal'] = macd_signal
    indicators['MACD_Hist'] = macd_diff

    # Bollinger Bands
    bb_high = ta.volatility.bollinger_hband(close, window=20)
    bb_mid = ta.volatility.bollinger_mavg(close, window=20)
    bb_low = ta.volatility.bollinger_lband(close, window=20)
    indicators['BB_Upper'] = bb_high
    indicators['BB_Middle'] = bb_mid
    indicators['BB_Lower'] = bb_low

    # On-Balance Volume (OBV)
    indicators['OBV'] = ta.volume.on_balance_volume(close, volume)

    return pd.DataFrame(indicators)

def get_market_sentiment(ticker):
    """Get market sentiment indicators"""
    stock = yf.Ticker(ticker)
    info = stock.info
    
    sentiment = {
        'Market Sentiment': {
            '52 Week High': info.get('fiftyTwoWeekHigh', 'N/A'),
            '52 Week Low': info.get('fiftyTwoWeekLow', 'N/A'),
            '50 Day Average': info.get('fiftyDayAverage', 'N/A'),
            '200 Day Average': info.get('twoHundredDayAverage', 'N/A'),
            'Short Ratio': info.get('shortRatio', 'N/A'),
            'Short Percent of Float': info.get('shortPercentOfFloat', 'N/A')
        }
    }
    return sentiment

def get_risk_metrics(ticker):
    """Calculate risk metrics"""
    stock = yf.Ticker(ticker)
    info = stock.info
    
    # Get historical data for volatility calculation
    hist = stock.history(period='1y')
    returns = hist['Close'].pct_change().dropna()
    
    risk_metrics = {
        'Risk Metrics': {
            'Beta': info.get('beta', 'N/A'),
            'Volatility (1Y)': f"{returns.std() * np.sqrt(252):.2%}",
            'Sharpe Ratio': f"{(returns.mean() * 252) / (returns.std() * np.sqrt(252)):.2f}",
            'Max Drawdown': f"{(hist['Close'].cummax() - hist['Close']).max() / hist['Close'].cummax():.2%}",
            'Value at Risk (95%)': f"{returns.quantile(0.05):.2%}"
        }
    }
    return risk_metrics

def create_stock_chart(data, technical_data, ticker):
    """Create an interactive stock chart with technical indicators"""
    # Create figure with secondary y-axis
    fig = make_subplots(rows=3, cols=1, 
                       shared_xaxes=True,
                       vertical_spacing=0.05,
                       row_heights=[0.6, 0.2, 0.2])

    # Add candlestick chart
    fig.add_trace(go.Candlestick(x=data.index,
                                open=data['Open'],
                                high=data['High'],
                                low=data['Low'],
                                close=data['Close'],
                                name='OHLC'),
                  row=1, col=1)

    # Add moving averages
    fig.add_trace(go.Scatter(x=data.index, y=technical_data['SMA_20'],
                            name='SMA 20',
                            line=dict(color='blue')),
                  row=1, col=1)
    fig.add_trace(go.Scatter(x=data.index, y=technical_data['SMA_50'],
                            name='SMA 50',
                            line=dict(color='orange')),
                  row=1, col=1)
    fig.add_trace(go.Scatter(x=data.index, y=technical_data['SMA_200'],
                            name='SMA 200',
                            line=dict(color='red')),
                  row=1, col=1)

    # Add Bollinger Bands
    fig.add_trace(go.Scatter(x=data.index, y=technical_data['BB_Upper'],
                            name='BB Upper',
                            line=dict(color='gray', dash='dash')),
                  row=1, col=1)
    fig.add_trace(go.Scatter(x=data.index, y=technical_data['BB_Lower'],
                            name='BB Lower',
                            line=dict(color='gray', dash='dash')),
                  row=1, col=1)

    # Add RSI
    fig.add_trace(go.Scatter(x=data.index, y=technical_data['RSI'],
                            name='RSI',
                            line=dict(color='purple')),
                  row=2, col=1)
    # Add RSI levels
    fig.add_hline(y=70, line_dash="dash", line_color="red", row=2, col=1)
    fig.add_hline(y=30, line_dash="dash", line_color="green", row=2, col=1)

    # Add MACD
    fig.add_trace(go.Scatter(x=data.index, y=technical_data['MACD'],
                            name='MACD',
                            line=dict(color='blue')),
                  row=3, col=1)
    fig.add_trace(go.Scatter(x=data.index, y=technical_data['MACD_Signal'],
                            name='Signal',
                            line=dict(color='orange')),
                  row=3, col=1)
    fig.add_trace(go.Bar(x=data.index, y=technical_data['MACD_Hist'],
                        name='Histogram',
                        marker_color='gray'),
                  row=3, col=1)

    # Update layout
    fig.update_layout(
        title=f'{ticker} Stock Price and Technical Indicators',
        yaxis_title='Price',
        yaxis2_title='RSI',
        yaxis3_title='MACD',
        xaxis_rangeslider_visible=False,
        height=800,
        template='plotly_white'
    )

    return fig

@app.route('/search_ticker', methods=['GET'])
def search_ticker():
    query = request.args.get('q', '')
    suggestions = get_ticker_suggestions(query)
    return jsonify(suggestions)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        ticker = request.form['ticker'].upper()
        start_date = request.form['start']
        end_date = request.form['end']
        interval = request.form['interval']
        
        # Download data
        data = yf.download(ticker, start=start_date, end=end_date, interval=interval)
        
        # Validate data
        if data is None or data.empty:
            error = f"No data found for ticker '{ticker}'. Please check the symbol and try again."
            return render_template('index.html', error=error)
        
        # Store price data in session for download
        session['price_data'] = data.reset_index().to_json(orient='split')
        session['ticker'] = ticker
        
        # Show price data table only
        price_table = data.reset_index().to_html(classes='table table-striped', index=False)
        return render_template('index.html', price_table=price_table, ticker=ticker)

    return render_template('index.html')

@app.route('/download_excel')
def download_excel():
    if 'price_data' not in session or 'ticker' not in session:
        return "No price data available to download.", 400
    import pandas as pd
    data = pd.read_json(session['price_data'], orient='split')
    ticker = session['ticker']
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        data.to_excel(writer, index=False, sheet_name='Price Data')
    output.seek(0)
    filename = f"{ticker}_price_data.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/download_report')
def download_report():
    if 'analysis_data' not in session:
        return "No analysis data available", 400
        
    data = session['analysis_data']
    ticker = data['ticker']
    start_date = data['start_date']
    end_date = data['end_date']
    interval = data['interval']
    required_return = data['required_return']
    growth_rate = data['growth_rate']
    
    # Download data
    stock_data = yf.download(ticker, start=start_date, end=end_date, interval=interval)
    
    # Calculate technical indicators
    technical_data = calculate_technical_indicators(stock_data)
    
    # Get financial ratios and valuation models
    ratios = get_financial_ratios(ticker)
    valuations = calculate_valuation_models(ticker, required_return, growth_rate)
    sentiment = get_market_sentiment(ticker)
    risk_metrics = get_risk_metrics(ticker)
    
    # Create Excel file with multiple sheets
    filename = f"{ticker}_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Sheet 1: Historical data
        stock_data.to_excel(writer, sheet_name='Historical Data')
        format_excel_worksheet(writer.sheets['Historical Data'], 
                             f'Historical Data - {ticker} ({start_date} to {end_date})')
        
        # Sheet 2: Technical Analysis
        technical_data.to_excel(writer, sheet_name='Technical Analysis')
        format_excel_worksheet(writer.sheets['Technical Analysis'],
                             f'Technical Analysis - {ticker}')
        
        # Sheet 3: Financial Analysis
        analysis_data = []
        
        # Add company info
        stock = yf.Ticker(ticker)
        info = stock.info
        analysis_data.extend([
            ['Company Information', '', ''],
            ['', 'Name', info.get('longName', 'N/A')],
            ['', 'Sector', info.get('sector', 'N/A')],
            ['', 'Industry', info.get('industry', 'N/A')],
            ['', 'Market Cap', f"${info.get('marketCap', 'N/A'):,.2f}" if info.get('marketCap') else 'N/A'],
            ['', 'Current Price', f"${info.get('currentPrice', 'N/A'):,.2f}" if info.get('currentPrice') else 'N/A'],
            ['', '', '']
        ])
        
        # Add financial ratios
        for category, metrics in ratios.items():
            analysis_data.append([category, '', ''])
            for metric, value in metrics.items():
                if isinstance(value, (int, float)):
                    formatted_value = f"{value:,.2f}"
                else:
                    formatted_value = value
                analysis_data.append(['', metric, formatted_value])
            analysis_data.append(['', '', ''])
        
        # Add market sentiment
        for category, metrics in sentiment.items():
            analysis_data.append([category, '', ''])
            for metric, value in metrics.items():
                if isinstance(value, (int, float)):
                    formatted_value = f"${value:,.2f}" if 'Price' in metric else f"{value:,.2f}"
                else:
                    formatted_value = value
                analysis_data.append(['', metric, formatted_value])
            analysis_data.append(['', '', ''])
        
        # Add risk metrics
        for category, metrics in risk_metrics.items():
            analysis_data.append([category, '', ''])
            for metric, value in metrics.items():
                analysis_data.append(['', metric, value])
            analysis_data.append(['', '', ''])
        
        # Add valuation models
        analysis_data.append(['Valuation Models', '', ''])
        for model, value in valuations.items():
            if isinstance(value, (int, float)):
                formatted_value = f"${value:,.2f}"
            else:
                formatted_value = value
            analysis_data.append(['', model, formatted_value])
        
        analysis_df = pd.DataFrame(analysis_data, columns=['Category', 'Metric', 'Value'])
        analysis_df.to_excel(writer, sheet_name='Financial Analysis', index=False)
        format_excel_worksheet(writer.sheets['Financial Analysis'], 
                             f'Financial Analysis - {ticker}')
    
    return send_file(filename, as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
